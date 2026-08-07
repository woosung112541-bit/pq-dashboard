import streamlit as st
import pandas as pd
import io
import json
import zipfile
import PyPDF2
import openpyxl
import google.generativeai as genai  
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload

# ==========================================
# ⚙️ [초기 세팅]
# ==========================================
st.set_page_config(page_title="PQ 자동화 대시보드", layout="wide")

if 'semi_fixed' not in st.session_state:
    st.session_state.semi_fixed = {
        "credit_rating": "BB-",       
        "new_hire_rate": 7.0,          
        "overlap_level": "최고 (350% 이상)",        
        "investment_ratio": 3.03,       
    }
if 'semi_fixed_confirmed' not in st.session_state: st.session_state.semi_fixed_confirmed = False
if 'dream_team' not in st.session_state: st.session_state.dream_team = {"pm": "", "pe": "", "pes": ""}
if 'notice_text' not in st.session_state: st.session_state.notice_text = ""
if 'raw_excel_bytes' not in st.session_state: st.session_state.raw_excel_bytes = b""
if 'final_excel_bytes' not in st.session_state: st.session_state.final_excel_bytes = b""
if 'db_summary_cache' not in st.session_state: st.session_state.db_summary_cache = {}

with st.sidebar:
    st.markdown("### 🧠 AI 엔진 설정")
    api_key = st.text_input("Gemini API Key를 입력하세요", type="password")
    if api_key:
        genai.configure(api_key=api_key)
        st.success("AI 엔진 연결 완료!")

# ==========================================
# 🔑 [Google Drive 연동 및 마스터 DB 로드]
# ==========================================
@st.cache_resource
def authenticate_google_drive():
    try:
        oauth_data = st.secrets["google_oauth"]
        creds = Credentials(
            token=None, refresh_token=oauth_data["refresh_token"],
            token_uri="https://oauth2.googleapis.com/token",
            client_id=oauth_data["client_id"], client_secret=oauth_data["client_secret"]
        )
        return build('drive', 'v3', credentials=creds)
    except Exception: return None

@st.cache_data(ttl=300)
def load_master_db_from_drive():
    try:
        drive_service = authenticate_google_drive()
        if not drive_service: return {}
        results = drive_service.files().list(q="name contains '마스터' and trashed=false", fields="files(id, mimeType)").execute()
        items = results.get('files', [])
        if not items: return {}
        
        file_id = items[0]['id']
        mime_type = items[0].get('mimeType', '')
        if mime_type == 'application/vnd.google-apps.spreadsheet':
            request = drive_service.files().export_media(fileId=file_id, mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        else:
            request = drive_service.files().get_media(fileId=file_id)
            
        fh = io.BytesIO()
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while not done: _, done = downloader.next_chunk()
        fh.seek(0)
        
        return pd.read_excel(fh, sheet_name=None)
    except Exception: return {}

def get_all_files_recursively(drive_service, folder_id, target_mime_types, depth=0):
    if depth > 5: return [] 
    found_files = []
    page_token = None
    while True:
        try:
            response = drive_service.files().list(q=f"'{folder_id}' in parents and trashed=false", fields='nextPageToken, files(id, name, mimeType)', pageToken=page_token).execute()
            for file in response.get('files', []):
                if file['mimeType'] in target_mime_types: found_files.append(file)
                elif file['mimeType'] == 'application/vnd.google-apps.folder':
                    found_files.extend(get_all_files_recursively(drive_service, file['id'], target_mime_types, depth + 1))
            page_token = response.get('nextPageToken', None)
            if not page_token: break
        except Exception: break
    return found_files

def get_all_subfolders_map(drive_service, root_id, depth=0):
    if depth > 5: return {}
    folder_dict = {}
    page_token = None
    while True:
        try:
            response = drive_service.files().list(q=f"'{root_id}' in parents and mimeType='application/vnd.google-apps.folder' and trashed=false", fields='nextPageToken, files(id, name)', pageToken=page_token).execute()
            for f in response.get('files', []):
                folder_dict[f['name']] = f['id']
                folder_dict.update(get_all_subfolders_map(drive_service, f['id'], depth + 1))
            page_token = response.get('nextPageToken', None)
            if not page_token: break
        except Exception: break
    return folder_dict

@st.cache_data(ttl=300)
def scan_drive_archive_cached():
    try:
        drive_service = authenticate_google_drive()
        if not drive_service: return {}
        res_arch = drive_service.files().list(q="name='기술인' and mimeType='application/vnd.google-apps.folder' and trashed=false", fields="files(id)").execute()
        if not res_arch.get('files'): return {}
        archive_id = res_arch.get('files')[0]['id']
        res_sub = drive_service.files().list(q=f"'{archive_id}' in parents and mimeType='application/vnd.google-apps.folder' and trashed=false", fields="files(id, name)").execute()
        archive_status = {}
        for folder in res_sub.get('files', []):
            all_pdfs = get_all_files_recursively(drive_service, folder['id'], ['application/pdf'])
            archive_status[folder['name']] = [f['name'] for f in all_pdfs]
        return archive_status
    except Exception: return {}

# 💡 [핵심 진화 1] 파이썬을 이용한 스마트 팩트 데이터 추출기 
def calculate_fact_data(master_db_dict):
    if not master_db_dict: return {}
    parsed_data = {}
    total_company_count = 0.0
    total_company_amt = 0.0
    
    for sheet_name, df in master_db_dict.items():
        if df.empty or str(sheet_name).startswith('Sheet'): continue
        name = str(sheet_name).strip()
        
        type_col = next((c for c in df.columns if any(k in str(c).replace(' ','') for k in ['공종', '분야', '시설물', '종류', '구분', '점검대상'])), None)
        days_col = next((c for c in df.columns if any(k in str(c).replace(' ','') for k in ['일수', '참여일', '기간', '인정'])), None)
        amt_col = next((c for c in df.columns if any(k in str(c).replace(' ','') for k in ['금액', '백만원', '준공', '도급'])), None)
        
        total_days, total_count, total_amt = 0.0, 0.0, 0.0
        
        for _, row in df.iterrows():
            if (days_col and pd.isna(row[days_col])) and (amt_col and pd.isna(row[amt_col])):
                continue
                
            weight = 0.8
            if type_col and pd.notna(row[type_col]):
                val_type = str(row[type_col]).replace(' ', '')
                if '교량' in val_type or '터널' in val_type: weight = 1.0
                
            d_val, a_val = 0, 0
            if days_col and pd.notna(row[days_col]):
                try: d_val = float(str(row[days_col]).replace(',', ''))
                except: pass
            total_days += d_val * weight
            
            if amt_col and pd.notna(row[amt_col]):
                try: a_val = float(str(row[amt_col]).replace(',', ''))
                except: pass
            total_amt += a_val * weight
            
            if d_val > 0 or a_val > 0:
                total_count += (1 * weight)
            
        years = total_days / 365.0
        if total_amt > 1000000: total_amt = total_amt / 1000000
        
        parsed_data[name] = {"years": round(years, 2), "count": round(total_count, 1), "amt": round(total_amt, 0)}
        total_company_count += total_count
        total_company_amt += total_amt
        
    parsed_data["COMPANY"] = {"count": round(total_company_count, 1), "amt": round(total_company_amt, 0)}
    return parsed_data

# ==========================================
# 🧠 [Backend Engine] 좌표 다이렉트 주입 엔진
# ==========================================
class PQScoringEngine:
    def parse_excel_structure(self, excel_bytes):
        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)
        ws = wb.active
        structure = []
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            row_str = " | ".join([str(c).strip() for c in row if c is not None and str(c).strip() != ""])
            if row_str:
                structure.append({"row_num": row_idx, "row_content": row_str})
        return structure

    def write_cell_safe(self, sheet, r, c, val):
        cell = sheet.cell(row=r, column=c)
        if type(cell).__name__ == 'MergedCell':
            for merged_range in sheet.merged_cells.ranges:
                if cell.coordinate in merged_range:
                    sheet.cell(row=merged_range.min_row, column=merged_range.min_col).value = val
                    break
        else:
            cell.value = val

    def run_ai_dreamteam_optimizer(self, notice_text, excel_bytes, semi_fixed, db_summary, dream_team):
        excel_structure = self.parse_excel_structure(excel_bytes)
        excel_json_str = json.dumps(excel_structure, ensure_ascii=False, indent=2)
        
        pm_data = db_summary.get(dream_team['pm'], {"years":0, "count":0, "amt":0})
        pe_data = db_summary.get(dream_team['pe'], {"years":0, "count":0, "amt":0})
        pes_data = db_summary.get(dream_team['pes'], {"years":0, "count":0, "amt":0})
        comp_data = db_summary.get("COMPANY", {"count":0, "amt":0})
        
        # 💡 [핵심 진화 2] AI에게 팩트를 줄 때, "어느 행에 어떤 데이터를 넣으라"고 지시.
        fact_instructions = f"""
        [적용해야 할 팩트 데이터 지시사항]
        - '사업책임' 경력/실적 관련 행에는 이 수치를 적용: {pm_data['years']}년, {pm_data['count']}건, {pm_data['amt']:,.0f} 백만원, 등급: 특급
        - '분야별책임' 경력/실적 관련 행에는 이 수치를 적용: {pe_data['years']}년, {pe_data['count']}건, {pe_data['amt']:,.0f} 백만원, 등급: 특급
        - '분야별참여' 경력/실적 관련 행에는 이 수치를 적용: {pes_data['years']}년, {pes_data['count']}건, {pes_data['amt']:,.0f} 백만원, 등급: 특급
        - '유사용역 수행실적' (회사 전체 실적) 관련 행에는 이 수치를 적용: {comp_data['count']}건, {comp_data['amt']:,.0f} 백만원
        
        - 신용평가등급: {semi_fixed.get('credit_rating')} -> 세부기준표에서 해당 구간 점수 적용 (BB-는 2.8점 등)
        - 업무중첩도: '{semi_fixed.get('overlap_level')}' -> 세부기준표에서 350%이상 최고 감점 구간 적용 (사책 3.6점, 분책 2.4점)
        - 신규고용율 가점: {semi_fixed.get('new_hire_rate')}% -> 세부기준 가점표 최고구간 (0.3점) 적용
        """
        
        prompt = f"""
        당신은 어떠한 수학적 연산도 하지 않는 '좌표 매칭 로봇'입니다.
        아래 [적용해야 할 팩트 데이터 지시사항]에 나열된 수치를 100% 맹신하여, [세부평가기준]의 배점 구간을 찾고 [자기평가표 엑셀 구조]의 정확한 행(row_num)에 <점수>와 <팩트>를 기입하세요.

        [★★★★★ 절대 준수 규칙]
        1. **계산 금지:** 스스로 숫자를 지어내거나 합치지 마세요.
        2. **지시사항 절대 복종:** 제가 [팩트 데이터 지시사항]에서 불러준 수치(예: 5.33년 등)를 그대로 'reason'칸에 적고 점수를 매기세요.
        3. **reason(산출근거) 초간결화:** 수식을 뺀 최종 숫자만 적으세요. (예: "5.33년", "22건", "1,288 백만원", "특급", "BB-", "350% 이상", "7.0%")
        4. **소계/총계 연산:** 엑셀 상의 '계', '소계', '총계' 행은 도출된 하위 항목 획득점수를 더해서 적고, 'reason'은 "-" 로 하세요.

        {fact_instructions}

        [세부평가기준 텍스트]
        {notice_text[:8000]} 

        [자기평가표 엑셀 구조 (이 번호에 맞춰 점수 기입)]
        {excel_json_str}

        오직 순수 JSON으로만 반환:
        {{
            "row_results": [
                {{"row_num": 1, "score": "<찾은 총점>", "reason": "-"}},
                {{"row_num": 8, "score": "<찾은 점수>", "reason": "<환산 경력>"}}
            ]
        }}
        """
        try:
            model = genai.GenerativeModel('gemini-3.6-flash')
            response = model.generate_content(prompt)
            result_text = response.text.strip().removeprefix("```json").removesuffix("```").strip()
            parsed = json.loads(result_text)
            
            ai_results = parsed.get("row_results", [])
            
            wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
            ws = wb.active
            
            score_col_idx, reason_col_idx = -1, -1
            for r_idx in range(1, min(6, ws.max_row + 1)):
                for c_idx, cell in enumerate(ws[r_idx], start=1):
                    val = str(cell.value).replace(' ', '').replace('\n','')
                    if '자기평가' in val or '점수' in val: score_col_idx = c_idx
                    elif '산출근거' in val or '산출' in val: reason_col_idx = c_idx
            
            if score_col_idx == -1: score_col_idx = ws.max_column - 1
            if reason_col_idx == -1: reason_col_idx = ws.max_column
            
            for res in ai_results:
                row_idx = res.get("row_num")
                if row_idx:
                    score = res.get("score", "")
                    reason = res.get("reason", "")
                    if score != "" and str(score).replace('.', '', 1).isdigit(): 
                        self.write_cell_safe(ws, row_idx, score_col_idx, float(score) if '.' in str(score) else int(score))
                    if reason != "" and reason != "점수": 
                        self.write_cell_safe(ws, row_idx, reason_col_idx, reason)
            
            out_bytes = io.BytesIO()
            wb.save(out_bytes)
            display_df = pd.DataFrame(ai_results)
            
            return out_bytes.getvalue(), display_df
            
        except Exception as e:
            st.error(f"연산 오류: {e}")
            return None, pd.DataFrame()

engine = PQScoringEngine()

# ==========================================
# 🖥️ [Frontend]
# ==========================================
st.title("PQ 자동화 대시보드")
st.caption("※ 시트 분류 & 수동 역할 배정 기반 환각 방탄 엔진")

tab1, tab2, tab3, tab4 = st.tabs(["📥 1. DB/서류 관리", "⚙️ 2. 인원 배정 및 팩트체크", "📊 3. 시뮬레이션", "🖨️ 4. 서류 패키징"])

with tab1:
    col1, col2 = st.columns(2)
    with col1:
        st.subheader("Zone A: 공고(PDF) & 자기평가표(Excel) 업로드")
        notice_files = st.file_uploader("", type=['pdf', 'xlsx'], accept_multiple_files=True)
        if notice_files and api_key and st.button("🧠 업로드 문서 시스템 적용", type="primary"):
            with st.spinner("엑셀 뼈대 구조를 고정하고 있습니다..."):
                notice_temp = ""
                excel_bytes_temp = b""
                for file in notice_files:
                    if file.name.lower().endswith('.pdf'):
                        pdf = PyPDF2.PdfReader(file)
                        for page in pdf.pages[:20]: notice_temp += page.extract_text() or ""
                    elif file.name.lower().endswith('.xlsx'):
                        excel_bytes_temp = file.getvalue()
                
                if excel_bytes_temp:
                    st.session_state.notice_text = notice_temp
                    st.session_state.raw_excel_bytes = excel_bytes_temp
                    st.success("✅ 문서 분석 완료 및 엑셀 템플릿 적용 완료.")
                else: st.error("엑셀 파일이 업로드되지 않았습니다.")
                
    with col2:
        st.subheader("Zone B: '기술인' 서류 스캔")
        if st.button("🔍 드라이브 재스캔"):
            scan_drive_archive_cached.clear()
            load_master_db_from_drive.clear()
            st.rerun()
        archive_data = scan_drive_archive_cached()
        if archive_data: st.success(f"📂 기술자 {len(archive_data)}명 스캔 완료!")

with tab2:
    st.markdown("### 📊 인원 배정 및 시스템 자동 계산 결과")
    st.caption("시스템이 마스터 DB의 각 시트를 스캔하여 환산 경력과 실적을 계산했습니다. 각 직책에 맞는 사람을 선택해 주세요.")
    
    db_dict = load_master_db_from_drive()
    if db_dict:
        # 데이터 계산 후 캐싱
        if not st.session_state.db_summary_cache:
            st.session_state.db_summary_cache = calculate_fact_data(db_dict)
            
        summary_data = st.session_state.db_summary_cache
        engineers = [name for name in summary_data.keys() if name != "COMPANY"]
        
        # 💡 [핵심 진화 3] AI가 맘대로 사람을 섞지 못하게 사용자가 드롭다운으로 명시적 지정
        st.write("#### 👤 투입 기술자 선택")
        col_pm, col_pe, col_pes = st.columns(3)
        with col_pm: 
            pm_name = st.selectbox("사업책임 (PM)", engineers, index=0 if len(engineers) > 0 else None)
        with col_pe: 
            pe_name = st.selectbox("분야별책임 (PE)", engineers, index=1 if len(engineers) > 1 else 0)
        with col_pes: 
            pes_name = st.selectbox("분야별참여 (PES)", engineers, index=2 if len(engineers) > 2 else 0)
            
        st.session_state.dream_team = {"pm": pm_name, "pe": pe_name, "pes": pes_name}
        
        st.write("#### 📈 파이썬 사전 계산 팩트 (이 수치가 AI에게 전달됩니다)")
        
        if pm_name:
            st.info(f"**[사업책임] {pm_name}:** 환산경력 {summary_data[pm_name]['years']}년 | 환산건수 {summary_data[pm_name]['count']}건 | 금액 {summary_data[pm_name]['amt']:,.0f}백만원")
        if pe_name:
            st.info(f"**[분야별책임] {pe_name}:** 환산경력 {summary_data[pe_name]['years']}년 | 환산건수 {summary_data[pe_name]['count']}건 | 금액 {summary_data[pe_name]['amt']:,.0f}백만원")
        if pes_name:
            st.info(f"**[분야별참여] {pes_name}:** 환산경력 {summary_data[pes_name]['years']}년 | 환산건수 {summary_data[pes_name]['count']}건 | 금액 {summary_data[pes_name]['amt']:,.0f}백만원")
        if "COMPANY" in summary_data:
            st.success(f"**[회사 전체 실적]:** 환산건수 {summary_data['COMPANY']['count']}건 | 금액 {summary_data['COMPANY']['amt']:,.0f}백만원")
            
    else:
        st.warning("구글 드라이브에 '마스터'가 포함된 엑셀 파일을 올려주세요.")

with tab3:
    st.markdown("### 🏆 점수 산출")
    
    with st.expander("🔒 팩트 데이터(반고정 항목) 적용 설정", expanded=not st.session_state.semi_fixed_confirmed):
        sf = st.session_state.semi_fixed
        col1, col2 = st.columns(2)
        with col1:
            sf['credit_rating'] = st.selectbox("신용평가등급 (예: BB-)", ["A-","BBB+","BBB0","BBB-","BB+","BB0","BB-","B+","B0","B-","CCC+"], index=6)
            sf['new_hire_rate'] = st.number_input("신규고용율 (%)", value=7.0, step=0.1)
        with col2:
            sf['overlap_level'] = st.selectbox("중첩도 (350% 이상)", ["최저", "중간", "최고 (350% 이상)"], index=2)
            sf['investment_ratio'] = st.number_input("투자비율 (%)", value=3.03, step=0.01)
        st.session_state.semi_fixed = sf
        st.session_state.semi_fixed_confirmed = st.checkbox("항목 확인 완료", value=st.session_state.semi_fixed_confirmed)

    if st.button("🚀 자기평가표 점수 자동 기입", type="primary", disabled=not st.session_state.semi_fixed_confirmed):
        if not st.session_state.raw_excel_bytes:
            st.error("엑셀 템플릿이 없습니다.")
        elif not st.session_state.db_summary_cache:
            st.error("마스터 DB 분석 결과가 없습니다. Tab 2를 확인하세요.")
        else:
            with st.spinner('시스템이 팩트 데이터를 엑셀 원본 파일의 좌표에 덮어쓰고 있습니다...'):
                final_excel, log_df = engine.run_ai_dreamteam_optimizer(
                    st.session_state.notice_text, 
                    st.session_state.raw_excel_bytes, 
                    st.session_state.semi_fixed, 
                    st.session_state.db_summary_cache,
                    st.session_state.dream_team
                )
                if final_excel:
                    st.success("🎉 산출 완료! Tab 4에서 완성된 원본 엑셀 파일을 다운로드하세요.")
                    st.session_state.final_excel_bytes = final_excel
                    st.dataframe(log_df, use_container_width=True)

with tab4:
    st.markdown("### 🖨️ 서류 출력 및 패키징")
    if st.session_state.dream_team.get("pm"):
        if st.button("🔄 증빙 서류 및 엑셀 결과 다운로드 (ZIP)", type="primary"):
            with st.spinner("압축 중..."):
                try:
                    drive_service = authenticate_google_drive()
                    arch_res = drive_service.files().list(q="name='기술인' and trashed=false", fields="files(id)").execute()
                    arch_id = arch_res['files'][0]['id']
                    f_map = get_all_subfolders_map(drive_service, arch_id)
                    
                    z_buf = io.BytesIO()
                    with zipfile.ZipFile(z_buf, "w") as z:
                        z.writestr("0_완성된_자기평가표(원본보존).xlsx", st.session_state.final_excel_bytes)
                        
                        dt = st.session_state.dream_team
                        for p_name in [dt['pm'], dt['pe'], dt['pes']]:
                            if p_name in f_map:
                                pdfs = get_all_files_recursively(drive_service, f_map[p_name], ['application/pdf'])
                                for p_file in pdfs:
                                    req = drive_service.files().get_media(fileId=p_file['id'])
                                    fh = io.BytesIO()
                                    dl = MediaIoBaseDownload(fh, req)
                                    while not dl.next_chunk()[1]: pass
                                    fh.seek(0)
                                    z.writestr(f"{p_name}/{p_file['name']}", fh.read())
                    st.download_button("📦 다운로드", data=z_buf.getvalue(), file_name="최종패키지.zip", mime="application/zip")
                except Exception as e: st.error(e)
